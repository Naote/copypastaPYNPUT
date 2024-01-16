import time
import clipboard
import pandas as pd
import re
from openpyxl import Workbook
from pynput.keyboard import Controller, Key
from pyperclip import PyperclipException
import json

wb = Workbook()
ws = wb.active

keyboard = Controller()

with open('./cities.json', 'r', encoding='utf-8') as f:
    cities_data = json.load(f)

df_naf = pd.read_csv("./naf2008-listes-completes-5-niveaux.csv", encoding="UTF-8", delimiter=None)
df_naf["id_5"] = df_naf["id_5"].str.replace(".", "")
valid_naf_ids = set(df_naf["id_5"].tolist())
valid_naf_labels = set(df_naf["label_5"].str.upper().tolist())
naf_mapping = dict(zip(df_naf["id_5"], df_naf["label_5"]))
communes = set(city['label'].upper() for city in cities_data['cities'])


def is_valid_data(col, data, previous_data=None):
    """Vérifie la valeur copié..."""
    data = data.strip()
    
    if col == 1:
        if not data or "." in data:
            print(f"Invalide, colonne {col}: données vide ou contient '.'")
            return False

    elif col == 2:
        if data == previous_data:
            print(f"Invalid, colonne {col}: Données identique à la colonne précédente.")
            return False    

    elif col == 3:
        address_keywords = ["RUE", "BOULEVARD", "AVENUE", "IMPASSE", "Bis", "Ter", "CHEMIN", "LOTISSEMENT", "COURS", 
                            "RESIDENCE", "ROUTE", "LIEU-DIT", "LIEU DIT", "PLACE", "ALLEE", "CH", "B", "C", "A", "CHEM", 
                            "VILLA", "ZI", "ZA", "QUARTIER", "MONTEE", "QUARTIER", "HAMEAU", "PARC", "HAM", "DOMAINE", 
                            "ZAC", "ZUP", "TRAVERSE", "AVE", "CARREFOUR", "DEPARTEMENTALE", "QUAI", "PASSAGE", 
                            "ZONE ARTISANALE", "ZONE INDUSTRIEL", "ZONE", "SQUARE", "VILLA", "R", "AV", "BLVD", "BLD", 
                            "CCIAL", "VOIE", "IMM", "IMMEUBLE", "BD", "RD-PT", "CIAL"]

        if not any(keyword in data for keyword in address_keywords):
            print(f"Invalide, colonne {col}: Adresse '{data}' ne contient aucun mot-clé")
            return False

    elif col == 4:
        if not data or not re.fullmatch(r'\d{5}', data):
            print(f"Invalide, colonne {col}: format code postal incorrect")
            return False

    elif col == 5:
        if not data or data.upper() not in communes:
            print(f"Invalide, colonne {col}: n'est pas une commune francaise ou introuvable")
            return False

    elif col in [6, 7, 8]:
        if not data:
            return True
        if not re.fullmatch(r'(\d{2} ){4}\d{2}', data):
            print(f"Invalide, colonne {col}: Num Tel format FR incorrecte")
            return False

    elif col == 9:
        if not data or not data.replace(" ", "").replace(",", "").isalpha():
            print(f"Invalide, colonne {col}: chaîne de caractère non alphabetic")
            return False

    elif col == 10:
        modified_data = data.replace(".", "")
        if not data or modified_data not in valid_naf_ids:
            print(f"Invalide, colonne {col}: code NAF invalide ou introuvable")
            return False

    elif col == 11:
        if not data:
            return True
        if not re.fullmatch(r'\d{14}', data):
            print(f"Invalide, colonne {col}: numéro de SIRET invalide")
            return False

    elif col == 12:
        valid_values = ["Inconnu", "0", "1 à 2", "3 à 5", "6 à 9"]
        if not data or (data not in valid_values and not re.fullmatch(r'\d+ à \d+', data)):
            print(f"Invalide, colonne {col}: données ne correspondent pas a la regex")
            return False

    elif col == 13:
        email_pattern = r"[a-zA-Z0-9._-]+@+[a-zA-Z0-9._-]+\.[a-zA-Z]{1,6}"
        if not data:
            return True
        if not re.search(email_pattern, data):
            print(f"Invalide, colonne {col}: E-mail incorrecte")
            return False
        
    return True




print("Placez le curseur à l'endroit où vous voulez commencer à copier-coller, puis appuyez sur Entrée.")
input("Appuyez sur Entrée lorsque vous êtes prêt...")

print("Le script commencera dans 5 secondes...")
time.sleep(5)

# Commencer à la première ligne de la feuille Excel
row = 1
col = 1

try:
    while True:
        clipboard_data = None
        valid_data_obtained = False
        retry_count = 0
        previous_column_data = None  
        
        while not valid_data_obtained and retry_count < 4:
            # Copier (Ctrl+C)
            with keyboard.pressed(Key.ctrl):
                keyboard.press('c')
                keyboard.release('c')
            time.sleep(0.0001)

            max_retries = 5
            for _ in range(max_retries):
                try:
                    clipboard_data = clipboard.paste().strip()
                    break
                except PyperclipException:
                    if _ < max_retries - 1:
                        time.sleep(0.1)
                    else:
                        raise

            print(f"Colonne {col}, ligne {row}, tentative {retry_count}: donnée capturée: '{clipboard_data}'")

            if is_valid_data(col, clipboard_data, previous_column_data):
                valid_data_obtained = True
            else:
                print(f"Données invalide : {col}, retry ({retry_count+1})...")
                retry_count += 1
                time.sleep(0.05)

        if not valid_data_obtained:
            print(f"Données invalide : {col} après plusieurs essais. passage a la colonne suivante.")
            keyboard.press(Key.right)
            keyboard.release(Key.right)
            col += 1
            continue

        if col == 10 and clipboard_data in valid_naf_ids:
            ws.cell(row=row, column=col-1, value=naf_mapping[clipboard_data])

        # Coller les données dans Excel
        ws.cell(row=row, column=col, value=clipboard_data)
        
        previous_column_data = clipboard_data
        
        col += 1
        time.sleep(0.0001)
        
        keyboard.press(Key.right)
        keyboard.release(Key.right)
        time.sleep(0.0001)
        
        if col == 15:
            print(f"colonne innatendu atteinte {col}, Row {row}. réinitialisation.")
            col = 1
            reset_count += 1
            continue
        
        if col == 14:
            keyboard.press(Key.down)
            keyboard.release(Key.down)
            col = 1
            row += 1
            reset_count = 0 
        if row % 1000 == 0:
            print(f"Sauvegarde périodique à la ligne {row}...")
            wb.save('backup_output_v5.xlsx')
        time.sleep(0.01)

except KeyboardInterrupt:
    print("Interruption. Sauvegarde...")
    wb.save('final_output_v5.xlsx')
    time.sleep(3)